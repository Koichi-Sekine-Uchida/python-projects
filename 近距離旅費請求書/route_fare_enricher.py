
# -*- coding: utf-8 -*-
"""
route_fare_enricher.py

使い方:
  1) 同一フォルダに CSV を置きます（または --csv パスを指定）
  2) APIキーを使う場合は、下記いずれかを環境変数に設定（.env に記載可）
     - SERPAPI_KEY（推奨）
     - GOOGLE_API_KEY + GOOGLE_CSE_ID（Programmable Search）
  3) 実行:
        python route_fare_enricher.py --csv your.csv
     引数未指定時は、フォルダ内最初の .csv を処理します。

動作:
  - CSV を Excel (xlsx) に変換し、1行目をヘッダとします。
  - 各行の L列 出発地、M列 経由地、N列 到着地 から検索クエリを生成。
  - Web検索で運賃らしき金額を抽出し、U列(検索結果金額)、V列(検索結果URL)に書き込みます。

注意:
  - Webサイトの仕様変更や利用規約により、結果が取得できない可能性があります。
  - 正確性が重要な用途では、結果の目視確認を推奨します。
"""
import os
import re
import json
import time
import argparse
import urllib.parse
from pathlib import Path
from typing import Optional, Tuple, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# --- 定数: Excel 列インデックス（1始まり） ---
COL_L = 12  # 出発地
COL_M = 13  # 経由地（空でもOK）
COL_N = 14  # 到着地
COL_U = 21  # 検索結果金額
COL_V = 22  # 検索結果URL

# --- 正規表現: 金額抽出（例: 1,230円 / 1230円 / ￥1,230 など） ---
YEN_PATTERNS = [
    re.compile(r'(?<!\d)(\d{1,3}(?:,\d{3})+|\d+)\s*円'),
    re.compile(r'￥\s*(\d{1,3}(?:,\d{3})+|\d+)\b'),
]

def parse_first_yen(text: str) -> Optional[str]:
    if not text:
        return None
    for pat in YEN_PATTERNS:
        m = pat.search(text)
        if m:
            amount = m.group(1)
            return f"{amount}円"
    return None

def build_query(frm: str, via: Optional[str], to: str) -> str:
    parts = [frm]
    if via and str(via).strip():
        parts.append(str(via).strip())
    parts.append(to)
    # 運賃/料金/片道 をキーワードに追加
    return " ".join(parts) + " 運賃 料金 片道"

def serpapi_search(query: str, api_key: str) -> Tuple[Optional[str], Optional[str]]:
    """SerpAPI(Google) で検索し、snippet/要約から金額、URL を推定"""
    import requests
    endpoint = "https://serpapi.com/search.json"
    params = {
        "engine": "google",
        "q": query,
        "hl": "ja",
        "num": "10",
        "api_key": api_key,
    }
    r = requests.get(endpoint, params=params, timeout=20)
    r.raise_for_status()
    data = r.json()
    # organic_results の snippet 等から金額を抽出
    for item in data.get("organic_results", []):
        snippet = item.get("snippet") or ""
        title = item.get("title") or ""
        link = item.get("link")
        cand = parse_first_yen(snippet) or parse_first_yen(title)
        if cand and link:
            return cand, link
    # news_results や answer_box 等にも一応あたる
    answer_box = data.get("answer_box") or {}
    if isinstance(answer_box, dict):
        ab_text = " ".join(str(answer_box.get(k) or "") for k in ["title", "answer", "snippet"])
        cand = parse_first_yen(ab_text)
        if cand:
            src = (answer_box.get("link") or 
                   (answer_box.get("list") or [{}])[0].get("link") if answer_box.get("list") else None)
            return cand, src
    return None, None

def google_cse_search(query: str, api_key: str, cse_id: str) -> Tuple[Optional[str], Optional[str]]:
    """Google Programmable Search Engine (CSE) で検索"""
    import requests
    endpoint = "https://www.googleapis.com/customsearch/v1"
    params = {"key": api_key, "cx": cse_id, "q": query, "hl": "ja", "num": 10}
    r = requests.get(endpoint, params=params, timeout=20)
    r.raise_for_status()
    data = r.json()
    for item in data.get("items", []):
        snippet = item.get("snippet") or ""
        title = item.get("title") or ""
        link = item.get("link")
        cand = parse_first_yen(snippet) or parse_first_yen(title)
        if cand and link:
            return cand, link
    return None, None

def naive_bing_scrape(query: str) -> Tuple[Optional[str], Optional[str]]:
    """APIなし fallback: Bing検索結果ページの簡易パース（精度低 & 変更に弱い）"""
    import requests
    headers = {
        "User-Agent": "Mozilla/5.0"
    }
    url = "https://www.bing.com/search?q=" + urllib.parse.quote(query)
    r = requests.get(url, headers=headers, timeout=20)
    r.raise_for_status()
    html = r.text
    # 非構造テキストから金額らしきものを拾う（弱いが最終手段）
    cand = parse_first_yen(html)
    # 最初の検索結果リンク候補を雑に取得
    m = re.search(r'<li class="b_algo".*?<h2><a href="([^"]+)"', html, flags=re.S)
    url = m.group(1) if m else None
    return cand, url

def find_csv_in_dir(dirpath: Path) -> Optional[Path]:
    csvs = sorted(dirpath.glob("*.csv"))
    return csvs[0] if csvs else None

def ensure_env():
    # .env があれば読み込む
    load_dotenv(override=False)

def write_dataframe_to_excel(df: pd.DataFrame, xlsx_path: Path, sheet_name: str = "Sheet1"):
    # pandas -> Excel
    df.to_excel(xlsx_path, index=False, sheet_name=sheet_name)

def enrich_with_fares(xlsx_path: Path, sheet_name: str = "Sheet1",
                      delay_sec: float = 2.0):
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    # ヘッダ行の U/V に見出しを設定（既存なら上書きしない）
    ws.cell(row=1, column=COL_U, value=ws.cell(row=1, column=COL_U).value or "検索結果金額")
    ws.cell(row=1, column=COL_V, value=ws.cell(row=1, column=COL_V).value or "検索結果URL")

    serp_key = os.getenv("SERPAPI_KEY")
    g_api = os.getenv("GOOGLE_API_KEY")
    g_cse = os.getenv("GOOGLE_CSE_ID")

    # 2行目以降を処理
    for row in range(2, ws.max_row + 1):
        frm = (ws.cell(row=row, column=COL_L).value or "").strip() if ws.cell(row=row, column=COL_L).value else ""
        via = (ws.cell(row=row, column=COL_M).value or "").strip() if ws.cell(row=row, column=COL_M).value else ""
        to  = (ws.cell(row=row, column=COL_N).value or "").strip() if ws.cell(row=row, column=COL_N).value else ""

        if not frm or not to:
            continue

        query = build_query(frm, via, to)

        amount = None
        url = None

        try:
            if serp_key:
                amount, url = serpapi_search(query, serp_key)
            elif g_api and g_cse:
                amount, url = google_cse_search(query, g_api, g_cse)
            else:
                amount, url = naive_bing_scrape(query)
        except Exception as e:
            # エラー時は空欄のまま
            amount, url = None, None

        if amount:
            ws.cell(row=row, column=COL_U, value=amount)
        if url:
            ws.cell(row=row, column=COL_V, value=url)

        # レート制限対策
        time.sleep(delay_sec)

    wb.save(xlsx_path)

def main():
    ensure_env()
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv", type=str, help="入力CSVパス（未指定時は同一フォルダの最初のCSVを使用）")
    parser.add_argument("--out", type=str, help="出力Excelパス（未指定時は CSV名.xlsx）")
    parser.add_argument("--delay", type=float, default=2.0, help="検索間隔(秒) デフォルト2.0")
    args = parser.parse_args()

    here = Path.cwd()
    csv_path = Path(args.csv) if args.csv else find_csv_in_dir(here)
    if not csv_path or not csv_path.exists():
        raise FileNotFoundError("CSVが見つかりません。同一フォルダにCSVを置くか、--csv で指定してください。")

    # CSV -> DataFrame
    # 1行目をヘッダとして読み込み、余計な空白はトリム
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    df.columns = [c.strip() for c in df.columns]

    # 出力先
    xlsx_path = Path(args.out) if args.out else csv_path.with_suffix(".xlsx")
    write_dataframe_to_excel(df, xlsx_path)

    # Web検索で運賃取得し、U/V列に追記
    enrich_with_fares(xlsx_path, delay_sec=args.delay)

    print(f"完了: {xlsx_path}")
    print("※ 正確性の担保のため、金額・URLは目視確認を推奨します。")
    if not (os.getenv("SERPAPI_KEY") or (os.getenv('GOOGLE_API_KEY') and os.getenv('GOOGLE_CSE_ID'))):
        print("注意: APIキーが未設定のため、簡易スクレイプFallbackで検索しています。精度が下がる可能性があります。")

if __name__ == "__main__":
    main()
