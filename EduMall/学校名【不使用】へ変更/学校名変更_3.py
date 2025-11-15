# -*- coding: utf-8 -*-
"""
学校名を「【不使用】」に更新するスクリプト（openpyxl 版・Selenium Manager）
- EdgeDriverは Selenium Manager が自動解決（Service/固定PATH不要）
- 設定.xlsx は openpyxl で読み取り（Excel COM / Sheet1 固定 依存を排除）
- 先頭シートから B1:URL / B2:ID / B3:PW / B4:学校名 / B5:顧客名 を優先して読む
- B1～B3 が空なら、「ヘッダ形式（ログインURL/ユーザー名/パスワード）」でも読み取る
- 画面の XPATH は、頂いたコードの流儀を踏襲（必要に応じて調整してください）
"""

import os
import sys
import time
import traceback
from typing import Tuple, Optional

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ====== パス ======
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "設定.xlsx")

# ====== 設定.xlsx を openpyxl で読む ======
def load_settings() -> Tuple[str, str, str, str, str]:
    """
    設定.xlsxの先頭シートから (url, user, pw, gakkou_name, kokyaku_name) を返す。
    - まず B1..B5（URL, ID, PW, 学校名, 顧客名）を試す
    - B1..B3 が空なら、1行目ヘッダの「ログインURL/login_url/url」「ユーザー名/username/id」
      「パスワード/password/pw」列から 2行目の値を取得する
    """
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"設定.xlsx が見つかりません: {EXCEL_PATH}")

    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("openpyxl が未インストールです。`pip install openpyxl` を実行してください。") from e

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sh = wb.worksheets[0]  # 先頭シート

    def sval(r, c) -> str:
        v = sh.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    url     = sval(1, 2)  # B1
    user    = sval(2, 2)  # B2
    pw      = sval(3, 2)  # B3
    gakkou  = sval(4, 2)  # B4 任意
    kokyaku = sval(5, 2)  # B5 任意

    # B1〜B3 が空の場合はヘッダ形式にも対応
    if not (url and user and pw):
        # 1行目ヘッダを取得
        headers = []
        for j in range(1, sh.max_column + 1):
            headers.append((sh.cell(row=1, column=j).value or ""))
        headers = [str(h).strip() for h in headers]

        def find_col(candidates) -> Optional[int]:
            lower = [h.lower() for h in headers]
            for cand in candidates:
                c = cand.lower()
                if c in lower:
                    return lower.index(c) + 1
            return None

        col_url = find_col(["ログインURL", "login_url", "url"])
        col_id  = find_col(["ユーザー名", "ユーザ", "username", "id", "アカウント", "account"])
        col_pw  = find_col(["パスワード", "password", "pw", "pwd"])

        if col_url and col_id and col_pw:
            # 2行目の値を想定
            url = (sh.cell(row=2, column=col_url).value or url) if url == "" else url
            user = (sh.cell(row=2, column=col_id).value or user) if user == "" else user
            pw = (sh.cell(row=2, column=col_pw).value or pw) if pw == "" else pw

            # 学校名/顧客名もあれば拾う（任意）
            col_g = find_col(["学校名", "gakkou", "school_name"])
            col_k = find_col(["顧客名", "customer", "kokyaku"])
            if col_g and not gakkou:
                gakkou = str(sh.cell(row=2, column=col_g).value or "").strip()
            if col_k and not kokyaku:
                kokyaku = str(sh.cell(row=2, column=col_k).value or "").strip()

    if not (url and user and pw):
        raise ValueError("設定.xlsx から URL/ID/PW を取得できませんでした。B1～B3 か ヘッダ形式を確認してください。")

    return str(url).strip(), str(user).strip(), str(pw).strip(), gakkou, kokyaku


# ====== Selenium（Selenium Manager を利用） ======
def build_driver(detach=True):
    edge_options = EdgeOptions()
    edge_options.use_chromium = True
    if detach:
        edge_options.add_experimental_option("detach", True)
    driver = webdriver.Edge(options=edge_options)  # ★ドライバ自動取得
    driver.maximize_window()
    return driver


# ====== 画面操作（頂いたXPATH流儀に合わせる） ======
def login(driver, url, user, pw):
    print("ログイン処理開始...")
    driver.get(url)
    time.sleep(5)
    driver.find_element(By.XPATH, "/html/body/div/div[2]/form/div[1]/input").send_keys(user)
    driver.find_element(By.XPATH, "/html/body/div/div[2]/form/div[2]/input").send_keys(pw)
    driver.find_element(By.XPATH, "/html/body/div/div[2]/form/div[3]/button").click()
    time.sleep(5)
    print("ログイン成功！")

def goto_school_list(driver):
    driver.get("https://school.edumall.jp/schl/CAAS11001")
    time.sleep(5)

def fill_filters(driver, gakkou_name, kokyaku_name):
    # 任意：空なら何もしない
    if gakkou_name:
        gakkou_name_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@data-bind='value:gakkoName']"))
        )
        driver.execute_script("""
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
        """, gakkou_name_input, gakkou_name)

    if kokyaku_name:
        kokyaku_name_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@data-bind='value:kokyakuName']"))
        )
        driver.execute_script("""
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
        """, kokyaku_name_input, kokyaku_name)

def click_search(driver):
    print("検索ボタンをクリックします...")
    search_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '検索')]"))
    )
    driver.execute_script("arguments[0].removeAttribute('disabled');", search_button)
    driver.execute_script("arguments[0].scrollIntoView(true);", search_button)
    time.sleep(1)
    driver.execute_script("arguments[0].click();", search_button)
    time.sleep(10)

def process_rows_on_current_page(driver):
    # 1ページ内ですべての行を処理（学校名の先頭に【不使用】を付与して更新）
    try:
        rows = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(
                (By.XPATH, "//*[@id='mainContens']/div[3]/div/datatable/div[2]/div/table/tbody/tr")
            )
        )
    except Exception as e:
        print(f"学校一覧の取得に失敗: {e}")
        return

    row_count = len(rows)
    print(f"このページの行数: {row_count}")

    for i in range(1, row_count + 1):
        print(f"【学校 {i} を処理】")
        # 学校IDクリックで詳細へ
        try:
            school_id_xpath = f"//*[@id='mainContens']/div[3]/div/datatable/div[2]/div/table/tbody/tr[{i}]/td[1]/span"
            school_id_elem = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, school_id_xpath)))
            driver.execute_script("arguments[0].scrollIntoView(true);", school_id_elem)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", school_id_elem)
            time.sleep(3)
        except Exception as e:
            print(f"行 {i} の学校IDクリック失敗: {e}")
            continue

        # 学校名の先頭に【不使用】を付与
        try:
            gakko_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='mainContens']/div[5]/div[2]/div[4]/div[1]/input"))
            )
            current_value = (gakko_input.get_attribute("value") or "")
            new_value = current_value if current_value.startswith("【不使用】") else ("【不使用】" + current_value)

            driver.execute_script("""
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            """, gakko_input, new_value)
            print(f"学校名を更新: {new_value}")
        except Exception as e:
            print(f"学校名入力の更新に失敗: {e}")
            goto_school_list(driver)
            continue

        # 更新ボタン
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
            update_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='mainContens']/div[12]/div[2]/button"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", update_button)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", update_button)
            print("更新ボタンをクリック")
            time.sleep(2)
        except Exception as e:
            print(f"更新ボタンクリック失敗: {e}")
            goto_school_list(driver)
            continue

        # ポップアップOK（出ないこともある）
        try:
            popup_ok_button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'OK')]"))
            )
            popup_ok_button.click()
            print("更新ポップアップOK")
            time.sleep(3)
        except Exception:
            pass

        # 一覧復帰確認
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[@id='mainContens']/div[3]/div/datatable/div[2]/div/table/tbody/tr")
                )
            )
            print("一覧に戻りました。")
            time.sleep(3)
        except Exception as e:
            print(f"一覧復帰確認に失敗: {e}")
            goto_school_list(driver)

def iterate_pages(driver):
    # ページャをざっくり解析
    try:
        spans = driver.find_elements(By.XPATH, "//a[@class='page-link']/span")
        nums = [int(s.text.strip()) for s in spans if s.text.strip().isdigit()]
        total = max(nums) if nums else 1
    except Exception as e:
        print(f"ページ数の取得に失敗: {e}")
        total = 1

    print(f"総ページ数: {total}")
    for p in range(1, total + 1):
        print(f"\n=== ページ {p} ===")
        if p > 1:
            try:
                page_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//a[@class='page-link']/span[text()='{p}']"))
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", page_link)
                time.sleep(1)
                driver.execute_script("arguments[0].click();", page_link)
                time.sleep(3)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//*[@id='mainContens']/div[3]/div/datatable/div[2]/div/table/tbody/tr")
                    )
                )
                time.sleep(3)
            except Exception as e:
                print(f"{p}ページ目への移動に失敗: {e}")
                break

        process_rows_on_current_page(driver)

# ====== Main ======
def main():
    driver = None
    try:
        url, user, pw, gakkou_name, kokyaku_name = load_settings()
        print(f"[INFO] URL/ID/PW を取得しました。学校名フィルタ='{gakkou_name}', 顧客名フィルタ='{kokyaku_name}'")

        driver = build_driver(detach=True)
        login(driver, url, user, pw)

        goto_school_list(driver)

        if gakkou_name or kokyaku_name:
            fill_filters(driver, gakkou_name, kokyaku_name)
            click_search(driver)

        iterate_pages(driver)
        print("[DONE] すべての更新処理が完了しました。")

    except Exception:
        print("[ERROR] 処理中に例外が発生しました。", file=sys.stderr)
        traceback.print_exc()
    finally:
        # 自動終了させたい場合は以下を有効化
        # if driver:
        #     try:
        #         driver.quit()
        #     except Exception:
        #         pass
        pass

if __name__ == "__main__":
    main()
