import openpyxl
import sys
import os

def split_excel_file_fast(input_file_path, rows_per_file):
    
    # 1. 入力ファイルパスの解析
    # ディレクトリとファイル名を分離
    input_dir = os.path.dirname(input_file_path)
    input_filename_full = os.path.basename(input_file_path)
    
    # 拡張子を除いたファイル名と拡張子を分離
    base_name, ext = os.path.splitext(input_filename_full)
    
    if ext.lower() not in ['.xlsx']:
        print(f"エラー: サポートされていないファイル形式 '{ext}' です。xlsxファイルを使用してください。")
        return

    # 2. 【高速化】読み取り専用モードで入力ファイルを開く
    try:
        # data_only=True: セルの値（数式の結果）を取得
        wb = openpyxl.load_workbook(input_file_path, read_only=True, data_only=True) 
        sheet = wb.active
    except FileNotFoundError:
        print(f"エラー: ファイル '{input_file_path}' が見つかりません。")
        return
    except Exception as e:
        print(f"エラー: ファイルの読み込み中に問題が発生しました。{e}")
        return
    
    # 3. 【高速化】ジェネレータでデータを一括取得し、ヘッダーとデータ行を分離
    rows_iterator = sheet.values

    try:
        header_row = next(rows_iterator)
    except StopIteration:
        print("エラー: Excelファイルにデータ行がありません。")
        wb.close()
        return

    # データ行をリストに格納
    data_rows = list(rows_iterator)
    total_data_rows = len(data_rows)
    
    # 分割ファイルの数を計算 (データ行のみで計算)
    num_files = (total_data_rows + rows_per_file - 1) // rows_per_file

    # 4. 分割ファイルを作成
    for i in range(num_files):
        # 【高速化】書き込み専用モードで新しいブックを作成
        new_wb = openpyxl.Workbook(write_only=True)
        new_sheet = new_wb.create_sheet()

        # ヘッダー行を書き込む
        new_sheet.append(header_row)

        # 分割範囲を計算 (データ行のインデックスで計算)
        start_index = i * rows_per_file
        end_index = min(start_index + rows_per_file, total_data_rows)
        
        # 分割データを一括で書き込み
        for row_data in data_rows[start_index:end_index]:
            new_sheet.append(row_data)

        # 5. 【修正点】元のファイル名に枝番を付けて保存
        # 例: RiyoshaInfo_児童生徒（【不使用】）_1.xlsx
        output_filename = f"{base_name}_{i+1}{ext}"
        
        # 入力ファイルと同じディレクトリに保存
        output_file_path = os.path.join(input_dir, output_filename)
        
        new_wb.save(output_file_path)
        print(f"分割ファイル '{output_file_path}' を保存しました。")

    # 入力ファイルを閉じる
    wb.close()

# 実行部分
rows_per_file = 10000

# コマンドライン引数チェック
if len(sys.argv) < 2:
    print("エラー: Excelファイルのパスを引数として指定してください。")
    print("使用例: python Excel-Split100.py \"/path/to/your/file.xlsx\"")
    # 実行を終了
    sys.exit(1)

# 【修正点】引数として渡されたファイルパスを使用
input_file_path_arg = sys.argv[1]

# Windowsパスでバックスラッシュを使用している可能性を考慮し、パスを正規化
# （今回はLinux/Mac風の引数指定なので不要かもしれませんが、汎用性のため残します）
if sys.platform.startswith('win'):
    input_file_path_arg = input_file_path_arg.replace('/', '\\')

# スクリプト実行
split_excel_file_fast(input_file_path_arg, rows_per_file)