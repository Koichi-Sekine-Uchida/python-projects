import openpyxl
import sys

def split_excel_file(input_file, rows_per_file):
    # 入力ファイルを開く
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # 全行数を取得
    total_rows = sheet.max_row

    # 分割ファイルの数を計算
    num_files = total_rows // rows_per_file
    if total_rows % rows_per_file != 0:
        num_files += 1

    # 分割ファイルを作成
    for i in range(num_files):
        # 新しいブックを作成
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        
         # 分割範囲を計算
        start_row = i * rows_per_file
        end_row = min(start_row + rows_per_file+1, total_rows)

        # 分割データをコピー
        for row in range(start_row + 1, end_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                new_sheet.cell(row=row - start_row, column=col).value = cell_value

        # 1行目を含めた分割ファイルに1行目を追加
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            new_sheet.cell(row=1, column=col).value = cell_value

        # 分割ファイルを保存
        output_file = f"{i+1}.xlsx"
        new_wb.save(output_file)
        print(f"分割ファイル '{output_file}' を保存しました。")

    # 入力ファイルを閉じる
    wb.close()

# 使用例
input_file = "C:\tools\python-projects\EXCEL分割\【不使用】\RiyoshaInfo_児童生徒（【不使用】）.xlsx"
input_file = './' + sys.argv[1]
rows_per_file = 100

split_excel_file(input_file, rows_per_file)
